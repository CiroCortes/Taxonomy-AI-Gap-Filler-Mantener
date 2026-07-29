import os
import shutil
import ctypes
import ctypes.wintypes
from decimal import Decimal
from django.core.management.base import BaseCommand
from taxonomy.models import SKUItem
import openpyxl


def copy_file_shared(src_path: str, dst_path: str):
    """Copia un archivo habilitando lectura/escritura compartida (Win32 API) por si está abierto en Excel."""
    if os.name == 'nt':
        kernel32 = ctypes.windll.kernel32
        GENERIC_READ = 0x80000000
        FILE_SHARE_READ = 0x00000001
        FILE_SHARE_WRITE = 0x00000002
        FILE_SHARE_DELETE = 0x00000004
        OPEN_EXISTING = 3
        FILE_ATTRIBUTE_NORMAL = 0x80

        handle = kernel32.CreateFileW(
            src_path,
            GENERIC_READ,
            FILE_SHARE_READ | FILE_SHARE_WRITE | FILE_SHARE_DELETE,
            None,
            OPEN_EXISTING,
            FILE_ATTRIBUTE_NORMAL,
            None
        )
        if handle != -1:
            buf = ctypes.create_string_buffer(1024 * 1024)
            bytes_read = ctypes.wintypes.DWORD()
            with open(dst_path, 'wb') as f_dst:
                while True:
                    success = kernel32.ReadFile(handle, buf, 1024 * 1024, ctypes.byref(bytes_read), None)
                    if not success or bytes_read.value == 0:
                        break
                    f_dst.write(buf.raw[:bytes_read.value])
            kernel32.CloseHandle(handle)
            return True
    shutil.copyfile(src_path, dst_path)
    return True


class Command(BaseCommand):
    help = "Carga el Maestro de Artículos SAP desde un archivo Excel a la base de datos."

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, required=True, help="Ruta al archivo Excel del Maestro de Artículos.")

    def handle(self, *args, **options):
        excel_path = options['file']
        if not os.path.isabs(excel_path):
            excel_path = os.path.abspath(excel_path)

        if not os.path.exists(excel_path):
            self.stderr.write(f"Error: El archivo {excel_path} no existe.")
            return

        self.stdout.write(f"Cargando archivo: {excel_path}...")
        
        # Intentar cargar directo, o usar copia temporal compartida si está bloqueado por Excel
        working_file = excel_path
        temp_copied = False
        try:
            wb = openpyxl.load_workbook(working_file, read_only=True)
        except PermissionError:
            self.stdout.write("Archivo en uso por otro proceso (Excel/OneDrive). Creando copia temporal...")
            temp_file = os.path.join(os.path.dirname(excel_path), "_temp_maestro_import.xlsx")
            copy_file_shared(excel_path, temp_file)
            working_file = temp_file
            temp_copied = True
            wb = openpyxl.load_workbook(working_file, read_only=True)

        sheet = wb.active
        rows_iter = sheet.iter_rows(values_only=True)
        
        headers = next(rows_iter, None)
        if not headers:
            self.stderr.write("El archivo Excel está vacío.")
            return

        # Normalizar headers a minusculas sin espacios
        header_map = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

        def get_val(row, col_name, default=None):
            idx = header_map.get(col_name)
            if idx is not None and idx < len(row):
                val = row[idx]
                return val if val is not None else default
            return default

        created_count = 0
        updated_count = 0
        incomplete_count = 0
        total_rows = 0

        self.stdout.write("Procesando filas de Excel...")

        from django.db import transaction

        # Pre-cargar SKUs existentes en un diccionario para evitar N+1 queries
        existing_skus = {s.item_code: s for s in SKUItem.objects.all()}

        items_to_create = []
        items_to_update = []

        for row in rows_iter:
            item_code = get_val(row, 'ItemCode')
            if not item_code:
                continue

            total_rows += 1
            item_code_str = str(item_code).strip()
            item_name_str = str(get_val(row, 'ItemName', '')).strip()

            stock_val = float(get_val(row, 'Stock', 0.0) or 0.0)
            costo_un_val = Decimal(str(get_val(row, '$Costo UN', 0.0) or 0.0))
            costo_tt_val = Decimal(str(get_val(row, '$Costo TT', 0.0) or 0.0))
            moneda_val = str(get_val(row, 'Moneda', '')) if get_val(row, 'Moneda') else None
            precio_lista_val = Decimal(str(get_val(row, 'Precio Lista', 0.0) or 0.0))

            cod_grupo_val = str(get_val(row, 'Cod.Grupo', '')) if get_val(row, 'Cod.Grupo') else None
            nombre_grupo_val = str(get_val(row, 'Nombre Grupo', '')) if get_val(row, 'Nombre Grupo') else None
            clase_val = str(get_val(row, 'Clase', '')) if get_val(row, 'Clase') else None
            familia_val = str(get_val(row, 'Familia', '')) if get_val(row, 'Familia') else None
            subfamilia_val = str(get_val(row, 'SubFamilia', '')) if get_val(row, 'SubFamilia') else None
            modelo_val = str(get_val(row, 'Modelo', '')) if get_val(row, 'Modelo') else None
            categoria_val = str(get_val(row, 'Categoria', '')) if get_val(row, 'Categoria') else None

            # Limpiar strings vacíos a None
            clase_val = clase_val if (clase_val and clase_val.strip()) else None
            familia_val = familia_val if (familia_val and familia_val.strip()) else None
            subfamilia_val = subfamilia_val if (subfamilia_val and subfamilia_val.strip()) else None
            categoria_val = categoria_val if (categoria_val and categoria_val.strip()) else None

            if item_code_str in existing_skus:
                sku_obj = existing_skus[item_code_str]
                sku_obj.item_name = item_name_str
                sku_obj.stock = stock_val
                sku_obj.costo_un = costo_un_val
                sku_obj.costo_tt = costo_tt_val
                sku_obj.moneda = moneda_val
                sku_obj.precio_lista = precio_lista_val
                sku_obj.cod_grupo = cod_grupo_val
                sku_obj.nombre_grupo = nombre_grupo_val
                sku_obj.clase = clase_val
                sku_obj.familia = familia_val
                sku_obj.subfamilia = subfamilia_val
                sku_obj.modelo = modelo_val
                sku_obj.categoria = categoria_val
                is_inc = sku_obj.check_incomplete()
                if is_inc:
                    incomplete_count += 1
                items_to_update.append(sku_obj)
                updated_count += 1
            else:
                sku_obj = SKUItem(
                    item_code=item_code_str,
                    item_name=item_name_str,
                    stock=stock_val,
                    costo_un=costo_un_val,
                    costo_tt=costo_tt_val,
                    moneda=moneda_val,
                    precio_lista=precio_lista_val,
                    cod_grupo=cod_grupo_val,
                    nombre_grupo=nombre_grupo_val,
                    clase=clase_val,
                    familia=familia_val,
                    subfamilia=subfamilia_val,
                    modelo=modelo_val,
                    categoria=categoria_val,
                )
                is_inc = sku_obj.check_incomplete()
                if is_inc:
                    incomplete_count += 1
                items_to_create.append(sku_obj)
                created_count += 1

        wb.close()
        if temp_copied and os.path.exists(working_file):
            try:
                os.remove(working_file)
            except Exception:
                pass

        self.stdout.write("Guardando registros en la base de datos en batch...")
        with transaction.atomic():
            if items_to_create:
                SKUItem.objects.bulk_create(items_to_create, batch_size=2000)
            if items_to_update:
                SKUItem.objects.bulk_update(items_to_update, fields=[
                    'item_name', 'stock', 'costo_un', 'costo_tt', 'moneda',
                    'precio_lista', 'cod_grupo', 'nombre_grupo', 'clase',
                    'familia', 'subfamilia', 'modelo', 'categoria',
                    'is_incomplete', 'pending_fields'
                ], batch_size=2000)

        self.stdout.write(self.style.SUCCESS(
            f"Proceso finalizado con éxito:\n"
            f"- Total filas procesadas: {total_rows}\n"
            f"- Registros creados: {created_count}\n"
            f"- Registros actualizados: {updated_count}\n"
            f"- Registros incompletos (pendientes auditoría IA): {incomplete_count}"
        ))
