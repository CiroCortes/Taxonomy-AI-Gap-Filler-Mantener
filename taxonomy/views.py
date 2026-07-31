import json
import os
from decimal import Decimal
from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.db import transaction
from django.utils import timezone
from taxonomy.models import SKUItem
from ai_engine.gap_classifier import GeminiGapClassifier
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()


def sku_list_view(request):
    """Vista principal del Dashboard del Maestro de Artículos PESCO."""
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')  # all, incomplete, complete, ai_processed
    missing_filter = request.GET.get('missing', '').strip()  # clase, familia, subfamilia, categoria

    # Filtros de taxonomía
    grupo_filter = request.GET.get('grupo', '').strip()
    clase_filter = request.GET.get('clase', '').strip()
    familia_filter = request.GET.get('familia', '').strip()
    subfamilia_filter = request.GET.get('subfamilia', '').strip()
    modelo_filter = request.GET.get('modelo', '').strip()
    categoria_filter = request.GET.get('categoria', '').strip()

    skus = SKUItem.objects.all().order_by('item_code')

    if query:
        skus = skus.filter(Q(item_code__icontains=query) | Q(item_name__icontains=query) | Q(nombre_grupo__icontains=query) | Q(modelo__icontains=query))

    if status_filter == 'incomplete':
        skus = skus.filter(is_incomplete=True)
    elif status_filter == 'complete':
        skus = skus.filter(is_incomplete=False)
    elif status_filter == 'ai_processed':
        skus = skus.filter(ai_processed=True)

    # Filtrar por campo específico pendiente
    if missing_filter == 'clase':
        skus = skus.filter(Q(clase__isnull=True) | Q(clase=''))
    elif missing_filter == 'familia':
        skus = skus.filter(Q(familia__isnull=True) | Q(familia=''))
    elif missing_filter == 'subfamilia':
        skus = skus.filter(Q(subfamilia__isnull=True) | Q(subfamilia=''))
    elif missing_filter == 'categoria':
        skus = skus.filter(Q(categoria__isnull=True) | Q(categoria=''))

    # Filtros por valor de atributo
    if grupo_filter:
        skus = skus.filter(nombre_grupo=grupo_filter)
    if clase_filter:
        skus = skus.filter(clase=clase_filter)
    if familia_filter:
        skus = skus.filter(familia=familia_filter)
    if subfamilia_filter:
        skus = skus.filter(subfamilia=subfamilia_filter)
    if modelo_filter:
        skus = skus.filter(modelo=modelo_filter)
    if categoria_filter:
        skus = skus.filter(categoria=categoria_filter)

    # Métricas KPI para el Dashboard
    total_count = SKUItem.objects.count()
    sin_clase_count = SKUItem.objects.filter(Q(clase__isnull=True) | Q(clase='')).count()
    sin_familia_count = SKUItem.objects.filter(Q(familia__isnull=True) | Q(familia='')).count()
    sin_subfamilia_count = SKUItem.objects.filter(Q(subfamilia__isnull=True) | Q(subfamilia='')).count()
    sin_categoria_count = SKUItem.objects.filter(Q(categoria__isnull=True) | Q(categoria='')).count()
    incomplete_total = SKUItem.objects.filter(is_incomplete=True).count()
    ai_processed_count = SKUItem.objects.filter(ai_processed=True).count()

    # Listas distintivas para los menús desplegables de filtro
    grupos_list = SKUItem.objects.exclude(nombre_grupo__isnull=True).exclude(nombre_grupo='').values_list('nombre_grupo', flat=True).distinct().order_by('nombre_grupo')
    clases_list = SKUItem.objects.exclude(clase__isnull=True).exclude(clase='').values_list('clase', flat=True).distinct().order_by('clase')
    familias_list = SKUItem.objects.exclude(familia__isnull=True).exclude(familia='').values_list('familia', flat=True).distinct().order_by('familia')
    subfamilias_list = SKUItem.objects.exclude(subfamilia__isnull=True).exclude(subfamilia='').values_list('subfamilia', flat=True).distinct().order_by('subfamilia')
    modelos_list = SKUItem.objects.exclude(modelo__isnull=True).exclude(modelo='').values_list('modelo', flat=True).distinct().order_by('modelo')[:100]
    categorias_list = SKUItem.objects.exclude(categoria__isnull=True).exclude(categoria='').values_list('categoria', flat=True).distinct().order_by('categoria')

    paginator = Paginator(skus, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
        'missing_filter': missing_filter,
        'grupo_filter': grupo_filter,
        'clase_filter': clase_filter,
        'familia_filter': familia_filter,
        'subfamilia_filter': subfamilia_filter,
        'modelo_filter': modelo_filter,
        'categoria_filter': categoria_filter,
        'total_count': total_count,
        'sin_clase_count': sin_clase_count,
        'sin_familia_count': sin_familia_count,
        'sin_subfamilia_count': sin_subfamilia_count,
        'sin_categoria_count': sin_categoria_count,
        'incomplete_total': incomplete_total,
        'ai_processed_count': ai_processed_count,
        'grupos_list': grupos_list,
        'clases_list': clases_list,
        'familias_list': familias_list,
        'subfamilias_list': subfamilias_list,
        'modelos_list': modelos_list,
        'categorias_list': categorias_list,
    }
    return render(request, 'taxonomy/sku_list.html', context)


def batch_ai_view(request):
    """Vista de la nueva página interactiva para Pruebas y Clasificación por Lotes con IA."""
    groups_summary = list(SKUItem.objects.values('nombre_grupo').annotate(
        total=Count('id'),
        pending=Count('id', filter=Q(is_incomplete=True))
    ).filter(pending__gt=0).order_by('-pending'))

    clases_list = SKUItem.objects.exclude(clase__isnull=True).exclude(clase='').values_list('clase', flat=True).distinct().order_by('clase')
    
    context = {
        'groups_summary': groups_summary,
        'clases_list': clases_list,
    }
    return render(request, 'taxonomy/batch_ai.html', context)


def process_batch_ai_ajax(request):
    """Vista AJAX para procesar un lote configurable de SKUs (ej: 20 SKUs) filtrados por grupo o clase."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    api_key = os.getenv('GEMINI_API_KEY') or os.getenv('GOOGLE_API_KEY')
    if not api_key or not api_key.strip():
        return JsonResponse({'success': False, 'error': 'No se encontró la clave GEMINI_API_KEY en tu archivo .env. Por favor agrégala en el archivo .env'}, status=400)

    try:
        data = json.loads(request.body)
        grupo_selected = data.get('grupo', '').strip()
        clase_selected = data.get('clase', '').strip()
        batch_limit = int(data.get('limit', 20))

        incomplete_qs = SKUItem.objects.filter(is_incomplete=True)

        if grupo_selected:
            incomplete_qs = incomplete_qs.filter(nombre_grupo=grupo_selected)
        if clase_selected:
            incomplete_qs = incomplete_qs.filter(clase=clase_selected)

        skus_to_process = list(incomplete_qs.order_by('item_code')[:batch_limit])
        if not skus_to_process:
            return JsonResponse({'success': False, 'error': 'No hay SKUs pendientes que coincidan con los filtros seleccionados.'}, status=400)

        classifier = GeminiGapClassifier(api_key=api_key)
        allowed_clases = list(SKUItem.objects.exclude(clase__isnull=True).exclude(clase='').values_list('clase', flat=True).distinct())
        allowed_familias = list(SKUItem.objects.exclude(familia__isnull=True).exclude(familia='').values_list('familia', flat=True).distinct())

        results = []

        for sku in skus_to_process:
            rag_matches = list(SKUItem.objects.filter(
                is_incomplete=False,
                nombre_grupo=sku.nombre_grupo
            ).values('item_name', 'clase', 'familia', 'subfamilia')[:3])

            current_data = {
                "clase": sku.clase,
                "familia": sku.familia,
                "subfamilia": sku.subfamilia,
                "categoria": sku.categoria
            }

            try:
                res = classifier.fill_missing_fields(
                    sku.item_name,
                    sku.nombre_grupo or "",
                    current_data,
                    sku.pending_fields,
                    rag_matches,
                    allowed_clases=allowed_clases,
                    allowed_familias=allowed_familias
                )

                if "clase" in sku.pending_fields and res.clase:
                    if not allowed_clases or res.clase in allowed_clases:
                        sku.clase = res.clase

                if "familia" in sku.pending_fields and res.familia:
                    if not allowed_familias or res.familia in allowed_familias:
                        sku.familia = res.familia

                if "subfamilia" in sku.pending_fields and res.subfamilia:
                    sku.subfamilia = res.subfamilia

                if "categoria" in sku.pending_fields and res.categoria:
                    sku.categoria = res.categoria

                sku.ai_confidence_score = res.confidence
                sku.ai_rationale = res.rationale
                sku.ai_processed = True
                sku.ai_processed_at = timezone.now()
                sku.check_incomplete()
                sku.save()

                results.append({
                    'pk': sku.pk,
                    'item_code': sku.item_code,
                    'item_name': sku.item_name,
                    'grupo': sku.nombre_grupo or '--',
                    'clase': sku.clase or '--',
                    'familia': sku.familia or '--',
                    'subfamilia': sku.subfamilia or '--',
                    'modelo': sku.modelo or '--',
                    'categoria': sku.categoria or '--',
                    'confidence': sku.ai_confidence_score,
                    'rationale': sku.ai_rationale,
                    'status': 'Completo' if not sku.is_incomplete else 'Incompleto'
                })
            except Exception as item_err:
                results.append({
                    'pk': sku.pk,
                    'item_code': sku.item_code,
                    'item_name': sku.item_name,
                    'grupo': sku.nombre_grupo or '--',
                    'error': str(item_err)
                })

        return JsonResponse({
            'success': True,
            'processed_count': len(results),
            'grupo': grupo_selected or 'Todos',
            'results': results
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def upload_sap_excel_view(request):
    """Vista AJAX para carga directa de archivo Excel del Maestro SAP con opción de reinicio a cero."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    if 'excel_file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No se adjuntó ningún archivo Excel.'}, status=400)

    excel_file = request.FILES['excel_file']
    reset_ai = request.POST.get('reset_ai', 'true') == 'true'
    clear_db = request.POST.get('clear_db', 'false') == 'true'

    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'error': 'El archivo debe tener formato .xlsx o .xls'}, status=400)

    try:
        if clear_db:
            SKUItem.objects.all().delete()

        wb = openpyxl.load_workbook(excel_file, read_only=True)
        sheet = wb.active
        rows_iter = sheet.iter_rows(values_only=True)

        headers = next(rows_iter, None)
        if not headers:
            return JsonResponse({'success': False, 'error': 'El archivo Excel está vacío.'}, status=400)

        header_map = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

        def get_val(row, col_name, default=None):
            idx = header_map.get(col_name)
            if idx is not None and idx < len(row):
                val = row[idx]
                return val if val is not None else default
            return default

        existing_skus = {s.item_code: s for s in SKUItem.objects.all()} if not clear_db else {}
        items_to_create = []
        items_to_update = []
        created_count = 0
        updated_count = 0
        incomplete_count = 0
        total_rows = 0

        for row in rows_iter:
            item_code = get_val(row, 'ItemCode')
            if not item_code:
                continue

            total_rows += 1
            item_code_str = str(item_code).strip()
            item_name_str = str(get_val(row, 'ItemName', '')).strip()

            stock_val = float(get_val(row, 'Stock', 0.0) or 0.0)
            costo_un_val = Decimal(str(get_val(row, '$Costo UN', 0.0) or 0.0))
            costo_tt_val = Decimal(str(get_val(row, '$Costo TT', 0.0) or 0.0))
            moneda_val = str(get_val(row, 'Moneda', '')) if get_val(row, 'Moneda') else None
            precio_lista_val = Decimal(str(get_val(row, 'Precio Lista', 0.0) or 0.0))

            cod_grupo_val = str(get_val(row, 'Cod.Grupo', '')) if get_val(row, 'Cod.Grupo') else None
            nombre_grupo_val = str(get_val(row, 'Nombre Grupo', '')) if get_val(row, 'Nombre Grupo') else None
            clase_val = str(get_val(row, 'Clase', '')) if get_val(row, 'Clase') else None
            familia_val = str(get_val(row, 'Familia', '')) if get_val(row, 'Familia') else None
            subfamilia_val = str(get_val(row, 'SubFamilia', '')) if get_val(row, 'SubFamilia') else None
            modelo_val = str(get_val(row, 'Modelo', '')) if get_val(row, 'Modelo') else None
            categoria_val = str(get_val(row, 'Categoria', '')) if get_val(row, 'Categoria') else None

            clase_val = clase_val if (clase_val and clase_val.strip()) else None
            familia_val = familia_val if (familia_val and familia_val.strip()) else None
            subfamilia_val = subfamilia_val if (subfamilia_val and subfamilia_val.strip()) else None
            modelo_val = modelo_val if (modelo_val and modelo_val.strip()) else None
            categoria_val = categoria_val if (categoria_val and categoria_val.strip()) else None

            if item_code_str in existing_skus:
                sku_obj = existing_skus[item_code_str]
                sku_obj.item_name = item_name_str
                sku_obj.stock = stock_val
                sku_obj.costo_un = costo_un_val
                sku_obj.costo_tt = costo_tt_val
                sku_obj.moneda = moneda_val
                sku_obj.precio_lista = precio_lista_val
                sku_obj.cod_grupo = cod_grupo_val
                sku_obj.nombre_grupo = nombre_grupo_val
                sku_obj.clase = clase_val
                sku_obj.familia = familia_val
                sku_obj.subfamilia = subfamilia_val
                sku_obj.modelo = modelo_val
                sku_obj.categoria = categoria_val

                if reset_ai:
                    sku_obj.ai_processed = False
                    sku_obj.ai_confidence_score = None
                    sku_obj.ai_rationale = None
                    sku_obj.ai_processed_at = None

                if sku_obj.check_incomplete():
                    incomplete_count += 1
                items_to_update.append(sku_obj)
                updated_count += 1
            else:
                sku_obj = SKUItem(
                    item_code=item_code_str,
                    item_name=item_name_str,
                    stock=stock_val,
                    costo_un=costo_un_val,
                    costo_tt=costo_tt_val,
                    moneda=moneda_val,
                    precio_lista=precio_lista_val,
                    cod_grupo=cod_grupo_val,
                    nombre_grupo=nombre_grupo_val,
                    clase=clase_val,
                    familia=familia_val,
                    subfamilia=subfamilia_val,
                    modelo=modelo_val,
                    categoria=categoria_val,
                )
                if sku_obj.check_incomplete():
                    incomplete_count += 1
                items_to_create.append(sku_obj)
                created_count += 1

        wb.close()

        with transaction.atomic():
            if items_to_create:
                SKUItem.objects.bulk_create(items_to_create, batch_size=2000)
            if items_to_update:
                SKUItem.objects.bulk_update(items_to_update, fields=[
                    'item_name', 'stock', 'costo_un', 'costo_tt', 'moneda',
                    'precio_lista', 'cod_grupo', 'nombre_grupo', 'clase',
                    'familia', 'subfamilia', 'modelo', 'categoria',
                    'is_incomplete', 'pending_fields', 'ai_processed',
                    'ai_confidence_score', 'ai_rationale', 'ai_processed_at'
                ], batch_size=2000)

        return JsonResponse({
            'success': True,
            'total_rows': total_rows,
            'created_count': created_count,
            'updated_count': updated_count,
            'incomplete_count': incomplete_count,
            'message': f"Maestro cargado con éxito. Procesadas {total_rows} filas ({created_count} creadas, {updated_count} actualizadas/reiniciadas)."
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def process_single_sku_ai(request, pk):
    """Vista AJAX para procesar la IA en un único SKU con restricciones estrictas de SAP."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    sku = get_object_or_404(SKUItem, pk=pk)
    api_key = os.getenv('GEMINI_API_KEY') or os.getenv('GOOGLE_API_KEY')

    if not api_key or not api_key.strip():
        return JsonResponse({'success': False, 'error': 'No se encontró la clave GEMINI_API_KEY en tu archivo .env. Agrega GEMINI_API_KEY=tu_clave en el archivo .env'}, status=400)

    try:
        classifier = GeminiGapClassifier(api_key=api_key)
        rag_matches = list(SKUItem.objects.filter(
            is_incomplete=False,
            nombre_grupo=sku.nombre_grupo
        ).values('item_name', 'clase', 'familia', 'subfamilia')[:3])

        current_data = {
            "clase": sku.clase,
            "familia": sku.familia,
            "subfamilia": sku.subfamilia,
            "categoria": sku.categoria
        }

        allowed_clases = list(SKUItem.objects.exclude(clase__isnull=True).exclude(clase='').values_list('clase', flat=True).distinct())
        allowed_familias = list(SKUItem.objects.exclude(familia__isnull=True).exclude(familia='').values_list('familia', flat=True).distinct())

        result = classifier.fill_missing_fields(
            sku.item_name,
            sku.nombre_grupo or "",
            current_data,
            sku.pending_fields,
            rag_matches,
            allowed_clases=allowed_clases,
            allowed_familias=allowed_familias
        )

        if "clase" in sku.pending_fields and result.clase:
            if not allowed_clases or result.clase in allowed_clases:
                sku.clase = result.clase

        if "familia" in sku.pending_fields and result.familia:
            if not allowed_familias or result.familia in allowed_familias:
                sku.familia = result.familia

        if "subfamilia" in sku.pending_fields and result.subfamilia:
            sku.subfamilia = result.subfamilia

        if "categoria" in sku.pending_fields and result.categoria:
            sku.categoria = result.categoria

        sku.ai_confidence_score = result.confidence
        sku.ai_rationale = result.rationale
        sku.ai_processed = True
        sku.ai_processed_at = timezone.now()
        sku.check_incomplete()
        sku.save()

        return JsonResponse({
            'success': True,
            'clase': sku.clase,
            'familia': sku.familia,
            'subfamilia': sku.subfamilia,
            'categoria': sku.categoria,
            'confidence': sku.ai_confidence_score,
            'rationale': sku.ai_rationale,
            'is_incomplete': sku.is_incomplete,
            'pending_fields': sku.pending_fields,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def update_sku_taxonomy(request, pk):
    """Vista AJAX para modificación manual de la taxonomía por parte del usuario."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    sku = get_object_or_404(SKUItem, pk=pk)
    try:
        data = json.loads(request.body)
        sku.clase = data.get('clase', sku.clase)
        sku.familia = data.get('familia', sku.familia)
        sku.subfamilia = data.get('subfamilia', sku.subfamilia)
        sku.modelo = data.get('modelo', sku.modelo)
        sku.categoria = data.get('categoria', sku.categoria)

        # Limpiar vacíos a None
        sku.clase = sku.clase.strip() if sku.clase and sku.clase.strip() else None
        sku.familia = sku.familia.strip() if sku.familia and sku.familia.strip() else None
        sku.subfamilia = sku.subfamilia.strip() if sku.subfamilia and sku.subfamilia.strip() else None
        sku.modelo = sku.modelo.strip() if sku.modelo and sku.modelo.strip() else None
        sku.categoria = sku.categoria.strip() if sku.categoria and sku.categoria.strip() else None

        sku.check_incomplete()
        sku.save()

        return JsonResponse({
            'success': True,
            'clase': sku.clase or '--',
            'familia': sku.familia or '--',
            'subfamilia': sku.subfamilia or '--',
            'modelo': sku.modelo or '--',
            'categoria': sku.categoria or '--',
            'is_incomplete': sku.is_incomplete,
            'pending_fields': sku.pending_fields,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def export_ti_excel_view(request):
    """Genera un archivo Excel profesional multi-pestaña para el equipo de TI (Formato Carga Masiva SAP DTW + Auditoría de Cambios)."""
    export_scope = request.GET.get('scope', 'ai_processed')  # ai_processed, all, incomplete, group
    grupo_filter = request.GET.get('grupo', '').strip()

    skus = SKUItem.objects.all().order_by('item_code')

    if grupo_filter:
        skus = skus.filter(nombre_grupo=grupo_filter)

    if export_scope == 'ai_processed':
        skus = skus.filter(ai_processed=True)
    elif export_scope == 'incomplete':
        skus = skus.filter(is_incomplete=True)

    wb = openpyxl.Workbook()

    # --- PESTAÑA 1: Carga Masiva SAP DTW ---
    ws1 = wb.active
    ws1.title = "Carga Masiva SAP DTW"

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill_blue = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')

    headers1 = [
        'ItemCode (SKU Clave SAP)', 'ItemName (Descripción)', 'Nombre Grupo',
        'Clase (U_Clase)', 'Familia (U_Familia)', 'SubFamilia (U_Marca)',
        'Modelo (U_Modelo)', 'Categoría (U_Categoria)', 'Estado Taxonomía', 'Acción Recomendada TI'
    ]

    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for sku in skus:
        action = "UPDATE OITM" if not sku.is_incomplete else "PENDIENTE COMPLETAR"
        ws1.append([
            sku.item_code,
            sku.item_name,
            sku.nombre_grupo or '',
            sku.clase or '',
            sku.familia or '',
            sku.subfamilia or '',
            sku.modelo or '',
            sku.categoria or '',
            'COMPLETO' if not sku.is_incomplete else 'INCOMPLETO',
            action
        ])

    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 55)

    # --- PESTAÑA 2: Auditoría Detallada IA / Usuario ---
    ws2 = wb.create_sheet(title="Auditoría de Evaluaciones IA")
    headers2 = [
        'ItemCode', 'ItemName', 'Grupo SAP', 'Clase', 'Familia',
        'SubFamilia (Marca)', 'Modelo', 'Categoría',
        'Confianza IA (%)', 'Razonamiento Técnico IA', 'Fecha Evaluación'
    ]
    ws2.append(headers2)

    header_fill_dark = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill_dark
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ai_skus = skus.filter(ai_processed=True)
    for sku in ai_skus:
        proc_date = sku.ai_processed_at.strftime('%Y-%m-%d %H:%M:%S') if sku.ai_processed_at else 'N/A'
        conf_pct = f"{int(sku.ai_confidence_score * 100)}%" if sku.ai_confidence_score else 'N/A'
        ws2.append([
            sku.item_code,
            sku.item_name,
            sku.nombre_grupo or '',
            sku.clase or '',
            sku.familia or '',
            sku.subfamilia or '',
            sku.modelo or '',
            sku.categoria or '',
            conf_pct,
            sku.ai_rationale or '',
            proc_date
        ])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    # --- PESTAÑA 3: Instrucciones de Carga TI ---
    ws3 = wb.create_sheet(title="Instrucciones Equipo TI")
    ws3.append(["INSTRUCCIONES PARA EL EQUIPO DE TI - ACTUALIZACIÓN MASIVA EN SAP ERP"])
    ws3.append([])
    ws3.append(["1. La pestaña 'Carga Masiva SAP DTW' contiene los códigos SKU con sus atributos taxonómicos finalizados."])
    ws3.append(["2. Los campos corresponden a las tablas estándar de SAP ERP: U_Clase, U_Familia, U_Marca, U_Modelo, U_Categoria."])
    ws3.append(["3. Para ejecutar la carga masiva mediante Data Transfer Workbench (DTW) o SQL Update:"])
    ws3.append(["   UPDATE OITM SET U_Clase = T.Clase, U_Familia = T.Familia, U_Marca = T.SubFamilia, U_Categoria = T.Categoria FROM OITM INNER JOIN TempTable T ON OITM.ItemCode = T.ItemCode"])
    ws3.append(["4. La pestaña 'Auditoría de Evaluaciones IA' contiene la trazabilidad y justificación técnica de cada asignación."])

    ws3['A1'].font = Font(name='Calibri', size=13, bold=True, color='1F4E78')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    group_str = f"_{grupo_filter.replace(' ', '_')}" if grupo_filter else ""
    filename = f"PESCO_Carga_Masiva_TI_SAP{group_str}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
